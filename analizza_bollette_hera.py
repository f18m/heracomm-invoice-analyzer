#!/usr/bin/env python3

import zipfile
import os
import re
import fitz  # PyMuPDF
import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, Reference
import argparse
import sys
from datetime import datetime, timedelta


class InvoiceAnalyzer:

    REGEX_PERIODO = r"Periodo: dal (\d{2}\.\d{2}\.\d{4}) al (\d{2}\.\d{2}\.\d{4})"
    INTESTAZIONE_BOLLETTA_ELETTRICA = "Bolletta energia elettrica"
    INTESTAZIONE_BOLLETTA_GAS = "Bolletta gas"

    def __init__(self, verbose: int = 0):
        self.verbose = verbose

    def __italian_number_to_float_safe(self, s: str) -> float:
        """Converte una stringa con numero in formato italiano (es. '1.234,56') in float"""
        try:
            return float(s.replace(".", "").replace(",", "."))
        except ValueError:
            if self.verbose > 0:
                print(f"⚠️ Attenzione: impossibile convertire '{s}' in float.")
            return None

    def __estrai_testo_delle_sotto_bollette(self, pdf_path: str) -> list[str]:
        """Estrae i dati richiesti da una singola bolletta PDF Hera e ritorna una lista
           di stringhe contenenti il contenuto di ogni sotto-bolletta identificata"""

        nome_file = os.path.basename(pdf_path)

        if self.verbose > 1:
            print("***")
        print(f"🔍 Inizio l'analisi di {pdf_path}...")

        sotto_bollette = []
        with fitz.open(pdf_path) as doc:
            text = ""
            for i in range(len(doc)):
                page_text = doc[i].get_text()

                # Se incontro intestazione gas → escludo
                if InvoiceAnalyzer.INTESTAZIONE_BOLLETTA_GAS in page_text:
                    if self.verbose > 1:
                        print(f"💬 Escludo pagina {i} con intestazione GAS in {nome_file}")
                    continue # skip
                elif InvoiceAnalyzer.INTESTAZIONE_BOLLETTA_ELETTRICA not in page_text:
                    if self.verbose > 1:
                        print(f"💬 Escludo pagina {i} con intestazione SCONOSCIUTA in {nome_file}")
                    continue # skip

                periodo_match = re.findall(InvoiceAnalyzer.REGEX_PERIODO, page_text)
                if len(periodo_match) == 1:
                    # trovato un periodo, è l'inizio di una nuova sotto-bolletta,
                    # salva il testo precedente (se esiste) come sotto-bolletta
                    if text:
                        sotto_bollette.append(text)
                        text = ""

                text += page_text

            if text:
                # salva l'ultima sotto-bolletta
                sotto_bollette.append(text)
                
        if self.verbose > 1:
            print(f"💬 Trovate {len(sotto_bollette)} sotto-bollette in {nome_file}")
            if self.verbose > 2:
                for i, sb in enumerate(sotto_bollette):
                    # scrivi il testo estratto in un file di debug
                    debug_file = pdf_path.replace(".pdf", f"_debug_{i + 1}.txt")
                    print(f"💬 Testo sotto-bolletta {i + 1} estratto nel file di debug: {debug_file}")
                    with open(debug_file, "w", encoding="utf-8") as f:
                        f.write(sb)

        return sotto_bollette

    def __estrai_dati_da_sotto_bolletta(self, pdf_path: str, text: str) -> dict:
        nome_file = os.path.basename(pdf_path)

        # Periodo (inizio e fine)
        periodo_match = re.findall(InvoiceAnalyzer.REGEX_PERIODO, text)
        if len(periodo_match) > 1:
            if self.verbose > 0:
                print(f"⚠️ Attenzione: trovati più periodi nella bolletta {nome_file}")
            return None  # Se troviamo più periodi, la bolletta non è valida
        elif len(periodo_match) == 1:
            periodo_inizio_str = periodo_match[0][0]
            periodo_fine_str = periodo_match[0][1]

            try:
                periodo_inizio = datetime.strptime(periodo_inizio_str, "%d.%m.%Y")
                periodo_fine = datetime.strptime(periodo_fine_str, "%d.%m.%Y")
            except ValueError:
                if self.verbose > 0:
                    print(f"⚠️ Attenzione: formato data non valido nella bolletta {nome_file}.")
                return None

        else:
            #periodo_inizio = periodo_fine = None
            if self.verbose > 0:
                print(f"⚠️ Attenzione: impossibile trovare il periodo nella bolletta {nome_file}.")
            return None  # Se non troviamo il periodo, la bolletta non è valida

        # Consumi per fasce e totale
        consumi_regex = [
            r"Consumo fatturato.*?([-\d,.]+)\s+([-\d,.]+)\s+([-\d,.]+)\s*kWh",
            # Alcune volte il formato è leggermente diverso... proviamo con una regex alternativa
            r"Consumo fatturato\s\(Chilowatt orari\)\n([-\d,.]+)\n([-\d,.]+)\n([-\d,.]+)\s*kWh"
        ]
        consumi_match = None
        for regex in consumi_regex:
            consumi_match = re.search(regex, text)
            if consumi_match:
                break

        if consumi_match:
            consumo_f1 = self.__italian_number_to_float_safe(consumi_match.group(1))
            consumo_f23 = self.__italian_number_to_float_safe(consumi_match.group(2))
            consumo_tot = self.__italian_number_to_float_safe(consumi_match.group(3))
        else:
            if self.verbose > 0:
                print(f"⚠️ Attenzione: impossibile trovare i consumi nella bolletta {nome_file}.")
            return None  # Se non troviamo i consumi, la bolletta non è valida

        # Totale energia elettrica (escludendo gas e altri servizi)
        elettricita_match = re.search(r"Totale bolletta/contratto\s+([\d,]+)", text)
        if elettricita_match:
            totale_elettricita = self.__italian_number_to_float_safe(elettricita_match.group(1))
        else:
            #totale_elettricita = None
            if self.verbose > 0:
                print(f"⚠️ Attenzione: impossibile trovare il totale energia nella bolletta {nome_file}.")
            return None  # Se non troviamo il totale, la bolletta non è valida

        if self.verbose > 1:
            print(f"💬 Bolletta {nome_file}: Periodo {periodo_inizio} - {periodo_fine}, Consumi F1={consumo_f1} kWh, F2+F3={consumo_f23} kWh, Totale={consumo_tot} kWh, Costo={totale_elettricita} €")

        return {
            "File": pdf_path, # string
            "Periodo Inizio": periodo_inizio, # datetime
            "Periodo Fine": periodo_fine, # datetime 
            "Consumo F1 (kWh)": consumo_f1, # float
            "Consumo F2+F3 (kWh)": consumo_f23, # float
            "Consumo Totale (kWh)": consumo_tot, # float
            "Totale Energia (€)": totale_elettricita, # float
        }


    def estrai_dati_bolletta(self, pdf_path: str) -> list[dict]:
        """Estrae i dati richiesti da una singola bolletta PDF Hera
           e ritorna per ogni sotto-bolletta identificata un dict coi campi estratti.
           Questa è l'unica API pubblica della classe."""

        # Ogni bolletta può essere composta da più sotto-bollette (es. luce + gas)
        # oppure (luce + luce)
        dati = []
        sotto_bollette = self.__estrai_testo_delle_sotto_bollette(pdf_path)
        for sotto_bol in sotto_bollette:
            dati_sotto_bol = self.__estrai_dati_da_sotto_bolletta(pdf_path, sotto_bol)
            if dati_sotto_bol:
                dati.append(dati_sotto_bol)
        return dati



class Tools:

    def crea_csv(self, dati_bollette: list[dict], csv_path: str) -> None:
        # Creazione DataFrame e salvataggio CSV
        df = pd.DataFrame(dati_bollette)
        df.to_csv(csv_path, index=False)
        print(f"✅ File CSV creato: {csv_path}")

    def crea_excel(self, dati_bollette: list[dict], excel_path: str) -> None:
        # Creazione DataFrame e salvataggio Excel
        df = pd.DataFrame(dati_bollette)
        df.to_excel(excel_path, index=False)
        print(f"✅ File Excel creato: {excel_path}")

    def aggiungi_grafici(self, excel_path: str) -> None:
        """Aggiunge grafici di consumi e costi all'Excel"""
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Creiamo una colonna con i periodi come etichette (Periodo Inizio - Periodo Fine)
        for i in range(2, ws.max_row + 1):
            periodo = f"{ws.cell(row=i, column=2).value} - {ws.cell(row=i, column=3).value}"
            ws.cell(row=i, column=8, value=periodo)

        cats = Reference(ws, min_col=8, min_row=2, max_row=ws.max_row)

        # === Grafico Consumi ===
        chart_consumi = LineChart()
        chart_consumi.title = "Andamento Consumi"
        chart_consumi.y_axis.title = "kWh"
        chart_consumi.x_axis.title = "Periodo"

        data_consumi = Reference(ws, min_col=6, min_row=1, max_row=ws.max_row)  # Consumo Totale
        chart_consumi.add_data(data_consumi, titles_from_data=True)
        chart_consumi.set_categories(cats)

        ws.add_chart(chart_consumi, "J2")

        # === Grafico Costi ===
        chart_costi = LineChart()
        chart_costi.title = "Andamento Costi Energia"
        chart_costi.y_axis.title = "Euro"
        chart_costi.x_axis.title = "Periodo"

        data_costi = Reference(ws, min_col=7, min_row=1, max_row=ws.max_row)  # Totale Energia
        chart_costi.add_data(data_costi, titles_from_data=True)
        chart_costi.set_categories(cats)

        ws.add_chart(chart_costi, "J20")

        wb.save(excel_path)
        print(f"📊 Grafici aggiunti a {excel_path}")

    def controlla_copertura(self, dati_sottobollette: list[dict]) -> list[tuple[datetime, datetime]]:
        """Verifica se ci sono buchi temporali tra le bollette"""

        df = pd.DataFrame(dati_sottobollette)
        df["Periodo Inizio"] = pd.to_datetime(df["Periodo Inizio"], format="%d.%m.%Y")
        df["Periodo Fine"] = pd.to_datetime(df["Periodo Fine"], format="%d.%m.%Y")

        # Ordinamento cronologico
        df = df.sort_values("Periodo Inizio").reset_index(drop=True)

        gaps = []
        for i in range(1, len(df)):
            prev_end = df.loc[i-1, "Periodo Fine"]
            curr_start = df.loc[i, "Periodo Inizio"]

            if curr_start > prev_end + timedelta(days=1):
                gaps.append((prev_end + timedelta(days=1), curr_start - timedelta(days=1)))

        return gaps

    def rinomina_pdfs(self, dati_sottobollette: list[dict]) -> None:
        """Rinomina i file PDF delle bollette in base ai dati estratti."""
        temp_dict = {}
        for dati in dati_sottobollette:
            curr_path = os.path.dirname(dati["File"])
            old_name = os.path.basename(dati["File"])
            periodo_inizio = dati["Periodo Inizio"].strftime("%Y%m%d")
            periodo_fine = dati["Periodo Fine"].strftime("%Y%m%d")
            anno = dati["Periodo Inizio"].year
            mese = dati["Periodo Inizio"].month
            nuovo_nome = f"{curr_path}/elettricita_{anno}_{mese:02}_{periodo_inizio}_{periodo_fine}.pdf"

            if dati["File"] in temp_dict:
                temp_dict[dati["File"]]["count"] += 1
            else:
                temp_dict[dati["File"]] = {"new_name":nuovo_nome, "count":1}

        for old_name, info in temp_dict.items():
            nuovo_nome = info["new_name"]
            if info["count"] > 1:
                base, ext = os.path.splitext(nuovo_nome)
                nuovo_nome = f"{base}_{info['count']}_sottobollette{ext}"

            if old_name != nuovo_nome:
                if os.path.exists(nuovo_nome):
                    print(f"⚠️ Impossibile rinominare {old_name} in {nuovo_nome}: il file di destinazione esiste già.")
                else:
                    os.rename(old_name, nuovo_nome)
                    print(f"🔄 Rinominato {old_name} in {nuovo_nome}")
            else:
                print(f"ℹ️ Il file {old_name} ha già il nome corretto.")



def main():
    parser = argparse.ArgumentParser(description="Estrai dati dalle bollette Hera e crea un Excel riepilogativo con grafici.")
    parser.add_argument("input_path", help="Percorso di un file ZIP di bollette o di una cartella contenente PDF")
    parser.add_argument("--output-csv", default="bollette_hera_riepilogo.csv", help="Nome del file CSV di output")
    parser.add_argument("--output-excel", default="bollette_hera_riepilogo.xlsx", help="Nome del file Excel di output")
    parser.add_argument("--verbose", type=int, help="Enable verbose output", default=0)
    parser.add_argument("--grafici", help="Aggiungi grafici nell'output", action='store_true')
    parser.add_argument("--rinomina",  help="Rinomina i files PDF con un formato human-friendly", action='store_true')
    args = parser.parse_args()

    input_path = args.input_path
    

    pdf_list = []

    if zipfile.is_zipfile(input_path):
        # Se è uno ZIP -> estraiamo i file
        extract_dir = "bollette_pdf"
        with zipfile.ZipFile(input_path, "r") as zip_ref:
            zip_ref.extractall(extract_dir)

        pdf_list = [os.path.join(extract_dir, f) for f in sorted(os.listdir(extract_dir)) if f.endswith(".pdf")]

    elif os.path.isdir(input_path):
        # Se è una cartella -> analizziamo tutti i files PDF al suo interno
        pdf_dir = input_path
        pdf_list = [os.path.join(pdf_dir, f) for f in sorted(os.listdir(pdf_dir)) if f.endswith(".pdf")]

    elif os.path.isfile(input_path) and input_path.endswith(".pdf"):
        # Se è un singolo file PDF, analizziamo solo quello
        pdf_list = [input_path]

    else:
        print("❌ Errore: devi fornire uno ZIP valido o una cartella contenente PDF.")
        sys.exit(1)

    # Elaborazione dei PDF
    x = InvoiceAnalyzer(verbose=args.verbose)
    dati_bollette = []
    print(f"✅ {len(pdf_list)} PDF files to analyze")
    for pdf_path in pdf_list:
        dati = x.estrai_dati_bolletta(pdf_path)
        if dati:
            for d in dati:
                dati_bollette.append(d)

    if not dati_bollette:
        print("❌ Nessun PDF analizzato correttamente.")
        sys.exit(1)

    t = Tools()

    # Rinomina dei PDF
    if args.rinomina:
        t.rinomina_pdfs(dati_bollette)

    if args.output_csv:
        t.crea_csv(dati_bollette, args.output_csv)

    if args.output_excel:
        t.crea_excel(dati_bollette, args.output_excel)
        # Aggiunta grafici
        if args.grafici:
            t.aggiungi_grafici(args.output_excel)

    if len(dati_bollette) > 1:
        buchi = t.controlla_copertura(dati_bollette)
        if buchi:
            print("⚠️ Trovati periodi non coperti:")
            for inizio, fine in buchi:
                print(f"   - dal {inizio.date()} al {fine.date()}")
        else:
            print("✅ Nessun buco temporale: le bollette coprono l'intero periodo senza interruzioni.")
       

if __name__ == "__main__":
    main()

