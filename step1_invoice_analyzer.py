#!/usr/bin/env python3

import zipfile
import os
import re
import fitz  # PyMuPDF
import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, Reference
import argparse
import sys
from datetime import datetime, timedelta


class InvoiceAnalyzer:

    INTESTAZIONE_BOLLETTA_ELETTRICA = "Bolletta energia elettrica"
    INTESTAZIONE_BOLLETTA_GAS = "Bolletta gas"
    MARKER_BOLLETTA_ELETTRICA = [
        "Bolletta energia elettrica",
        "Energia elettrica",
        "Scontrino dell’energia",
        "Scontrino dell'energia",
    ]

    FORMATO_LEGACY = "legacy"
    FORMATO_2026 = "hera_2026"

    REGEX_PERIODO = {
        FORMATO_LEGACY: r"Periodo:\s*dal\s*(\d{2}\.\d{2}\.\d{4})\s*al\s*(\d{2}\.\d{2}\.\d{4})",
        FORMATO_2026: r"Periodo\s+oggetto\s+di\s+fatturazione:\s*dal\s*(\d{2}\.\d{2}\.\d{4})\s*al\s*(\d{2}\.\d{2}\.\d{4})",
    }

    REGEX_SPESE_IN_EURO = {
        FORMATO_LEGACY: {
            "materia_energia": r"Spesa per la materia energia\s+([-\d.,]+)\s*€",
            "trasporto_e_contatore": r"Spesa per il trasporto e la gestione del contatore\s+([-\d.,]+)\s*€",
            "oneri_di_sistema": r"Spesa per oneri di sistema\s+([-\d.,]+)\s*€",
            "imposte_e_iva": r"Totale imposte e IVA\s+([-\d.,]+)\s*€",
            "totale_bolletta": r"Totale bolletta/contratto\s+([-\d.,]+)\s*€",
        },
        FORMATO_2026: {
            "materia_energia": r"Quota per consumi\s+[-\d.,]+\s*kWh\s+([-\d.,]+)\s*€",
            "trasporto_e_contatore": r"Quota fissa e quota potenza\s+[-\d.,]+\s*mesi\s+([-\d.,]+)\s*€",
            "oneri_di_sistema": r"[-\d.,]+\s*kW\s+per\s+[-\d.,]+\s*mesi\s+([-\d.,]+)\s*€",
            "imposte_e_iva": r"Accise e IVA\s+([-\d.,]+)\s*€",
            "totale_bolletta": r"Totale bolletta\s+([-\d.,]+)\s*€",
        },
    }

    REGEX_CONSUMI_IN_KWH = {
        FORMATO_LEGACY: [
            r"Consumo fatturato.*?([-\d.,]+)\s+([-\d.,]+)\s+([-\d.,]+)\s*kWh",
            # Alcune volte il formato è leggermente diverso... proviamo con una regex alternativa
            r"Consumo fatturato\s*\(Chilowatt\s+orari\)\s*([-\d.,]+)\s*([-\d.,]+)\s*([-\d.,]+)\s*kWh",
        ],
        FORMATO_2026: [
            r"Consumo fatturato\s*\(Chilowatt\s*orari\)\s*([-\d.,]+)\s*([-\d.,]+)\s*([-\d.,]+)",
            r"F1\s*\(kWh\)\s*F2\+F3\s*\(kWh\)\s*Totale\s*\(kWh\).*?\(\d+\s+giorni\)\s*([-\d.,]+)\s*([-\d.,]+)\s*([-\d.,]+)",
        ],
    }

    def __init__(self, verbose: int = 0, dump_debug: bool = False):
        self.verbose = verbose
        self.dump_debug = dump_debug

    def __italian_number_to_float_safe(self, s: str) -> float:
        """Converte una stringa con numero in formato italiano (es. '1.234,56') in float"""
        try:
            return float(s.replace(".", "").replace(",", "."))
        except ValueError:
            if self.verbose > 0:
                print(f"⚠️ Attenzione: impossibile convertire '{s}' in float.")
            return None

    def __detect_pdf_format(self, text: str) -> str:
        """Rileva automaticamente il formato del PDF in base ai marker testuali."""
        if re.search(r"Periodo\s+oggetto\s+di\s+fatturazione", text, flags=re.IGNORECASE):
            return InvoiceAnalyzer.FORMATO_2026
        return InvoiceAnalyzer.FORMATO_LEGACY

    def __search_first_match(self, regex_list: list[str], text: str):
        for regex in regex_list:
            m = re.search(regex, text, flags=re.IGNORECASE | re.DOTALL)
            if m:
                return m
        return None

    def __extract_spesa_with_fallback(self, text: str, voce_spesa: str, formato: str) -> float:
        # Prima prova il formato rilevato, poi i regex degli altri formati come fallback.
        ordered_formats = [formato] + [f for f in InvoiceAnalyzer.REGEX_SPESE_IN_EURO.keys() if f != formato]
        for fmt in ordered_formats:
            regex = InvoiceAnalyzer.REGEX_SPESE_IN_EURO[fmt][voce_spesa]
            match = re.search(regex, text, flags=re.IGNORECASE | re.DOTALL)
            if match:
                return self.__italian_number_to_float_safe(match.group(1))
        return 0.0

    def __estrai_testo_delle_sotto_bollette(self, pdf_path: str) -> list[str]:
        """Estrae i dati richiesti da una singola bolletta PDF Hera e ritorna una lista
           di stringhe contenenti il contenuto di ogni sotto-bolletta identificata"""

        nome_file = os.path.basename(pdf_path)

        if self.verbose > 1:
            print("***")
        print(f"🔍 Inizio l'analisi di {pdf_path}...")

        sotto_bollette = []
        with fitz.open(pdf_path) as doc:
            text = ""
            collecting_electric_bill = False
            for i in range(len(doc)):
                page_text = doc[i].get_text()
                has_period_marker = self.__search_first_match(list(InvoiceAnalyzer.REGEX_PERIODO.values()), page_text) is not None
                has_electricity_marker = any(marker in page_text for marker in InvoiceAnalyzer.MARKER_BOLLETTA_ELETTRICA)

                # Se incontro intestazione gas → escludo
                if InvoiceAnalyzer.INTESTAZIONE_BOLLETTA_GAS in page_text:
                    if self.verbose > 1:
                        print(f"💬 Escludo pagina {i} con intestazione GAS in {nome_file}")
                    continue # skip

                # Attiva la raccolta quando troviamo marker della bolletta elettrica.
                if has_electricity_marker or has_period_marker:
                    collecting_electric_bill = True

                if not collecting_electric_bill:
                    if self.verbose > 1:
                        print(f"💬 Escludo pagina {i} con intestazione SCONOSCIUTA in {nome_file}")
                    continue # skip

                if has_period_marker:
                    # trovato un periodo, è l'inizio di una nuova sotto-bolletta,
                    # salva il testo precedente (se esiste) come sotto-bolletta
                    if text:
                        sotto_bollette.append(text)
                        text = ""

                text += page_text

            if text:
                # salva l'ultima sotto-bolletta
                sotto_bollette.append(text)
                
        if self.verbose > 1:
            print(f"💬 Trovate {len(sotto_bollette)} sotto-bollette in {nome_file}")

        if self.dump_debug:
            for i, sb in enumerate(sotto_bollette):
                # scrivi il testo estratto in un file di debug
                debug_file = pdf_path.replace(".pdf", f"_debug_{i + 1}.txt")
                print(f"💬 Testo sotto-bolletta {i + 1} estratto nel file di debug: {debug_file}")
                with open(debug_file, "w", encoding="utf-8") as f:
                    f.write(sb)

        return sotto_bollette

    def __estrai_dati_da_sotto_bolletta(self, pdf_path: str, text: str) -> dict:
        nome_file = os.path.basename(pdf_path)
        formato = self.__detect_pdf_format(text)

        if self.verbose > 1:
            print(f"💬 Formato PDF rilevato per {nome_file}: {formato}")

        # Periodo (inizio e fine)
        periodo_match = re.search(InvoiceAnalyzer.REGEX_PERIODO[formato], text, flags=re.IGNORECASE | re.DOTALL)
        if not periodo_match:
            # Fallback: prova i regex periodo degli altri formati.
            periodo_match = self.__search_first_match(list(InvoiceAnalyzer.REGEX_PERIODO.values()), text)

        if periodo_match:
            periodo_inizio_str = periodo_match.group(1)
            periodo_fine_str = periodo_match.group(2)

            try:
                periodo_inizio = datetime.strptime(periodo_inizio_str, "%d.%m.%Y")
                periodo_fine = datetime.strptime(periodo_fine_str, "%d.%m.%Y")
            except ValueError:
                if self.verbose > 0:
                    print(f"⚠️ Attenzione: formato data non valido nella bolletta {nome_file}.")
                return None

            numero_giorni = (periodo_fine - periodo_inizio).days + 1
            if numero_giorni < 1:
                if self.verbose > 0:
                    print(f"⚠️ Attenzione: periodo non valido ({periodo_inizio_str} - {periodo_fine_str}) nella bolletta {nome_file}.")
                return None

        else:
            #periodo_inizio = periodo_fine = None
            if self.verbose > 0:
                print(f"⚠️ Attenzione: impossibile trovare il periodo nella bolletta {nome_file}.")
            return None  # Se non troviamo il periodo, la bolletta non è valida

        # Consumi per fasce e totale
        consumi_match = self.__search_first_match(InvoiceAnalyzer.REGEX_CONSUMI_IN_KWH[formato], text)
        if not consumi_match:
            # Fallback: prova anche i regex consumi degli altri formati.
            for fmt, regexes in InvoiceAnalyzer.REGEX_CONSUMI_IN_KWH.items():
                if fmt == formato:
                    continue
                consumi_match = self.__search_first_match(regexes, text)
                if consumi_match:
                    break

        if consumi_match:
            consumo_f1 = self.__italian_number_to_float_safe(consumi_match.group(1))
            consumo_f23 = self.__italian_number_to_float_safe(consumi_match.group(2))
            consumo_tot = self.__italian_number_to_float_safe(consumi_match.group(3))
        else:
            if self.verbose > 0:
                print(f"⚠️ Attenzione: impossibile trovare i consumi nella bolletta {nome_file}.")
            return None  # Se non troviamo i consumi, la bolletta non è valida

        # # Totale energia elettrica (escludendo gas e altri servizi)
        # elettricita_match = re.search(r"Totale bolletta/contratto\s+([\d,]+)", text)
        # if elettricita_match:
        #     totale_elettricita = self.__italian_number_to_float_safe(elettricita_match.group(1))
        # else:
        #     #totale_elettricita = None
        #     if self.verbose > 0:
        #         print(f"⚠️ Attenzione: impossibile trovare il totale energia nella bolletta {nome_file}.")
        #     return None  # Se non troviamo il totale, la bolletta non è valida

        # Voci di spesa
        voci_spesa = {}
        for voce_spesa in InvoiceAnalyzer.REGEX_SPESE_IN_EURO[formato].keys():
            voci_spesa[voce_spesa] = self.__extract_spesa_with_fallback(text, voce_spesa, formato)

        # Fine estrazione
        if self.verbose > 1:
            print(f"💬 Bolletta {nome_file}: Periodo {periodo_inizio} - {periodo_fine} ({numero_giorni} giorni), Consumi F1={consumo_f1}kWh, F2+F3={consumo_f23}kWh, Totale={consumo_tot}kWh, Costo={voci_spesa['totale_bolletta']}€")

        return {
            "file": pdf_path, # string
            "periodo_inizio": periodo_inizio, # datetime
            "periodo_fine": periodo_fine, # datetime 
            "numero_giorni": numero_giorni, # int
            "consumo_f1_kwh": consumo_f1, # float
            "consumo_f23_kwh": consumo_f23, # float
            "consumo_totale_kwh": consumo_tot, # float
            "materia_energia_eur": voci_spesa['materia_energia'], # float
            "trasporto_e_contatore_eur": voci_spesa['trasporto_e_contatore'], # float
            "oneri_di_sistema_eur": voci_spesa['oneri_di_sistema'], # float
            "imposte_e_iva_eur": voci_spesa['imposte_e_iva'], # float
            "totale_bolletta_eur": voci_spesa['totale_bolletta'], # float
        }

    def estrai_dati_bolletta(self, pdf_path: str) -> list[dict]:
        """Estrae i dati richiesti da una singola bolletta PDF Hera
           e ritorna per ogni sotto-bolletta identificata un dict coi campi estratti.
           Questa è l'unica API pubblica della classe."""

        # Ogni bolletta può essere composta da più sotto-bollette (es. luce + gas)
        # oppure (luce + luce)
        dati = []
        sotto_bollette = self.__estrai_testo_delle_sotto_bollette(pdf_path)
        for sotto_bol in sotto_bollette:
            dati_sotto_bol = self.__estrai_dati_da_sotto_bolletta(pdf_path, sotto_bol)
            if dati_sotto_bol:
                dati.append(dati_sotto_bol)
        return dati



class Tools:

    def __init__(self, dati_bollette: list[dict], verbose: int = 0):
        self.dati_bollette = dati_bollette
        self.verbose = verbose

    def crea_csv(self, csv_path: str) -> None:
        # Creazione DataFrame e salvataggio CSV
        df = pd.DataFrame(self.dati_bollette)
        df.to_csv(csv_path, index=False)
        print(f"✅ File CSV creato: {csv_path}")

    def crea_excel(self, excel_path: str) -> None:
        # Creazione DataFrame e salvataggio Excel
        df = pd.DataFrame(self.dati_bollette)
        df.to_excel(excel_path, index=False)
        print(f"✅ File Excel creato: {excel_path}")

    def aggiungi_grafici(self, excel_path: str) -> None:
        """Aggiunge grafici di consumi e costi all'Excel"""
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Creiamo una colonna con i periodi come etichette (periodo_inizio - periodo_fine)
        for i in range(2, ws.max_row + 1):
            periodo = f"{ws.cell(row=i, column=2).value} - {ws.cell(row=i, column=3).value}"
            ws.cell(row=i, column=8, value=periodo)

        cats = Reference(ws, min_col=8, min_row=2, max_row=ws.max_row)

        # === Grafico Consumi ===
        chart_consumi = LineChart()
        chart_consumi.title = "Andamento Consumi"
        chart_consumi.y_axis.title = "kWh"
        chart_consumi.x_axis.title = "Periodo"

        data_consumi = Reference(ws, min_col=6, min_row=1, max_row=ws.max_row)  # Consumo Totale
        chart_consumi.add_data(data_consumi, titles_from_data=True)
        chart_consumi.set_categories(cats)

        ws.add_chart(chart_consumi, "J2")

        # === Grafico Costi ===
        chart_costi = LineChart()
        chart_costi.title = "Andamento Costi Energia"
        chart_costi.y_axis.title = "Euro"
        chart_costi.x_axis.title = "Periodo"

        data_costi = Reference(ws, min_col=7, min_row=1, max_row=ws.max_row)  # Totale Energia
        chart_costi.add_data(data_costi, titles_from_data=True)
        chart_costi.set_categories(cats)

        ws.add_chart(chart_costi, "J20")

        wb.save(excel_path)
        print(f"📊 Grafici aggiunti a {excel_path}")

    def controlla_copertura(self) -> list[tuple[datetime, datetime]]:
        """Verifica se ci sono buchi temporali tra le bollette"""

        df = pd.DataFrame(self.dati_bollette)
        df["periodo_inizio"] = pd.to_datetime(df["periodo_inizio"], format="%d.%m.%Y")
        df["periodo_fine"] = pd.to_datetime(df["periodo_fine"], format="%d.%m.%Y")

        # Ordinamento cronologico
        df = df.sort_values("periodo_inizio").reset_index(drop=True)

        gaps = []
        for i in range(1, len(df)):
            prev_end = df.loc[i-1, "periodo_fine"]
            curr_start = df.loc[i, "periodo_inizio"]

            if curr_start > prev_end + timedelta(days=1):
                gaps.append((prev_end + timedelta(days=1), curr_start - timedelta(days=1)))

        return gaps

    def rinomina_pdfs(self) -> None:
        """Rinomina i file PDF delle bollette in base ai dati estratti."""
        temp_dict = {}

        def normalize_path(path: str) -> str:
            return os.path.normcase(os.path.normpath(path))

        for dati in self.dati_bollette:
            source_file = dati.get("file") or dati.get("File")
            if not source_file:
                raise KeyError("Missing 'file' key in extracted invoice data")

            curr_path = os.path.dirname(source_file)
            periodo_inizio = dati["periodo_inizio"].strftime("%Y%m%d")
            periodo_fine = dati["periodo_fine"].strftime("%Y%m%d")
            anno = dati["periodo_inizio"].year
            mese = dati["periodo_inizio"].month
            nuovo_nome = os.path.join(curr_path, f"elettricita_{anno}_{mese:02}_{periodo_inizio}_{periodo_fine}.pdf")

            if source_file in temp_dict:
                temp_dict[source_file]["count"] += 1
            else:
                temp_dict[source_file] = {"new_name":nuovo_nome, "count":1}

        for old_name, info in temp_dict.items():
            nuovo_nome = info["new_name"]
            if info["count"] > 1:
                base, ext = os.path.splitext(nuovo_nome)
                nuovo_nome = f"{base}_{info['count']}_sottobollette{ext}"

            old_normalized = normalize_path(old_name)
            new_normalized = normalize_path(nuovo_nome)

            if old_normalized != new_normalized:
                if os.path.exists(nuovo_nome):
                    print(f"⚠️ Impossibile rinominare {old_name} in {nuovo_nome}: il file di destinazione esiste già.")
                else:
                    os.rename(old_name, nuovo_nome)
                    print(f"🔄 Rinominato {old_name} in {nuovo_nome}")
            else:
                print(f"ℹ️ Il file {old_name} ha già il nome corretto.")

    def genera_sommario(self, summary_type: str, summary_format: str) -> None:
        """Genera un sommario testuale delle bollette analizzate."""
        if summary_type == "detailed":
            # Stampa un sommario testuale
            df = pd.DataFrame(self.dati_bollette)
            if "file" in df.columns:
                df["nome_file_pdf"] = df["file"].apply(os.path.basename)
            df = df.sort_values("periodo_inizio").reset_index(drop=True)
            print("\n📄 Sommario Bollette:")
            columns = ["nome_file_pdf", "periodo_inizio", "periodo_fine", "consumo_totale_kwh", "totale_bolletta_eur", "numero_giorni"]
            if summary_format == "html":
                print(df[columns].to_html(index=False))
            else:
                print(df[columns].to_string(index=False))
        elif summary_type == "yearly":
            print("⚠️ Avviso: il sommario annuale è stato disabilitato in questo step. Usa lo step2_interpolazione.py per un sommario accurato.")
        # il sommario annuale implementato sotto è molto IMPRECISO a causa delle bollette
        # che possono coprire periodi che attraversano più anni.
        # Lo step2 di interpolazione creerà un sommario più accurato.
        # elif summary_type == "yearly":
        #     # Stampa un sommario annuale
        #     df = pd.DataFrame(self.dati_bollette)
        #     df["Anno"] = df["periodo_inizio"].dt.year
        #     summary = df.groupby("Anno").agg({
        #         "Consumo Totale (kWh)": "sum",
        #         "Totale Energia (€)": "sum",
        #         "Numero Giorni": "sum"
        #     }).reset_index()

        #     print("\n📄 Sommario Annuale Bollette:")
        #     if summary_format == "html":
        #         print(summary.to_html(index=False))
        #     else:
        #         print(summary.to_string(index=False))
    
def main():
    parser = argparse.ArgumentParser(description="Estrai dati dalle bollette Hera e crea un Excel riepilogativo con grafici.")
    parser.add_argument("input_path", help="Percorso di un file ZIP di bollette o di una cartella contenente PDF")
    parser.add_argument("--output-csv", default="bollette_hera_riepilogo.csv", help="Nome del file CSV di output")
    parser.add_argument("--output-excel", default="", help="Nome del file Excel di output")
    parser.add_argument("--output-summary", default="detailed", help="Scrivi in output un sommario su base annuale, o più dettagliata", choices=["detailed", "yearly", "none"])
    parser.add_argument("--summary-format", default="text", help="Formato del sommario", choices=["text", "html"])
    parser.add_argument("--verbose", type=int, help="Enable verbose output", default=0)
    parser.add_argument("--dump-debug", help="Salva i testi estratti delle sotto-bollette in file TXT di debug", action='store_true')
    parser.add_argument("--grafici", help="Aggiungi grafici nell'output", action='store_true')
    parser.add_argument("--rinomina",  help="Rinomina i files PDF con un formato human-friendly", action='store_true')
    args = parser.parse_args()

    input_path = args.input_path
    

    pdf_list = []

    if zipfile.is_zipfile(input_path):
        # Se è uno ZIP -> estraiamo i file
        extract_dir = "bollette_pdf"
        with zipfile.ZipFile(input_path, "r") as zip_ref:
            zip_ref.extractall(extract_dir)

        pdf_list = [os.path.join(extract_dir, f) for f in sorted(os.listdir(extract_dir)) if f.endswith(".pdf")]

    elif os.path.isdir(input_path):
        # Se è una cartella -> analizziamo tutti i files PDF al suo interno
        pdf_dir = input_path
        pdf_list = [os.path.join(pdf_dir, f) for f in sorted(os.listdir(pdf_dir)) if f.endswith(".pdf")]

    elif os.path.isfile(input_path) and input_path.endswith(".pdf"):
        # Se è un singolo file PDF, analizziamo solo quello
        pdf_list = [input_path]

    else:
        print("❌ Errore: devi fornire uno ZIP valido o una cartella contenente PDF.")
        sys.exit(1)

    # Elaborazione dei PDF
    x = InvoiceAnalyzer(verbose=args.verbose, dump_debug=args.dump_debug)
    dati_bollette = []
    pdf_falliti = []
    print(f"✅ {len(pdf_list)} PDF files to analyze")
    for pdf_path in pdf_list:
        try:
            dati = x.estrai_dati_bolletta(pdf_path)
        except Exception as e:
            print(f"❌ Errore durante l'analisi del PDF {pdf_path}: {e}")
            pdf_falliti.append(pdf_path)
            continue

        if dati:
            for d in dati:
                dati_bollette.append(d)
        else:
            print(f"❌ Nessun dato estratto da {pdf_path}: nessuna sotto-bolletta elettrica valida trovata.")
            pdf_falliti.append(pdf_path)

    if not dati_bollette:
        print("❌ Nessun PDF analizzato correttamente.")
        sys.exit(1)

    if pdf_falliti:
        print(f"⚠️ PDF senza dati validi: {len(pdf_falliti)} su {len(pdf_list)}")

    t = Tools(dati_bollette)

    # Rinomina dei PDF
    if args.rinomina:
        t.rinomina_pdfs()

    if args.output_csv:
        t.crea_csv(args.output_csv)

    if args.output_excel:
        t.crea_excel(args.output_excel)
        # Aggiunta grafici
        if args.grafici:
            t.aggiungi_grafici(args.output_excel)

    if len(dati_bollette) > 1:
        buchi = t.controlla_copertura()
        if buchi:
            print("⚠️ Trovati periodi non coperti:")
            for inizio, fine in buchi:
                print(f"   - dal {inizio.date()} al {fine.date()}")
        else:
            print("✅ Nessun buco temporale: le bollette coprono l'intero periodo senza interruzioni.")

    if args.output_summary != "none":
        t.genera_sommario(args.output_summary, args.summary_format)

if __name__ == "__main__":
    main()

